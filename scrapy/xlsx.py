from openpyxl import Workbook


class WorkbookModule():

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active

    def save(self, file_name):
        self.wb.save(file_name)

    def create_sheet(self, sheet_name):
        self.wb.create_sheet(sheet_name)
