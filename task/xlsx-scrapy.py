from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup


def scrapy_google(company_name):
    result = requests.get(f'https://www.google.com/search?q={company_name}/')
    soup = BeautifulSoup(result.text, 'html.parser')
    data = soup.findAll(True, {'class': 'r'})
    return data


def main():
    # エクセルファイルのロード
    excel_path = 'C:/sample/sample.xlsm'
    workbook = load_workbook(filename=excel_path, read_only=True)

    # シートのロード
    sheet = workbook['企業情報']

    # セルの値取得
    search_company_data = []
    for index, row in enumerate(sheet.rows):
        if row[0].row == 1:
            # １行目はリターン
            return
        else:
            company_name = row[index].col[0]
            data = scrapy_google(company_name)
            search_company_data += data

    sheet = workbook.create_sheet('検索結果')
    for i, result in enumerate(search_company_data):
        sheet.row[i].col[0] = result

    workbook.close()
